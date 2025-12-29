import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import csv
import shutil
from collections import Counter
from datetime import datetime, timedelta
import random
import string
import sys

# macOS için zbar kütüphane yolu
if sys.platform == 'darwin':
    os.environ['DYLD_LIBRARY_PATH'] = '/opt/homebrew/lib:' + os.environ.get('DYLD_LIBRARY_PATH', '')

# Opsiyonel kütüphaneler
try:
    from openpyxl import load_workbook
    EXCEL_DESTEGI = True
except ImportError:
    EXCEL_DESTEGI = False

try:
    import barcode
    from barcode.writer import ImageWriter
    BARKOD_OLUSTURMA_DESTEGI = True
except ImportError:
    BARKOD_OLUSTURMA_DESTEGI = False

try:
    import cv2
    from pyzbar import pyzbar
    BARKOD_OKUMA_DESTEGI = True
except ImportError:
    BARKOD_OKUMA_DESTEGI = False

try:
    from PIL import Image, ImageTk
    PIL_DESTEGI = True
except ImportError:
    PIL_DESTEGI = False

try:
    import customtkinter as ctk
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    CTK_DESTEGI = True
except ImportError:
    CTK_DESTEGI = False

class GirisEkrani:
    """Öğretmen ve öğrenci giriş ekranı - Modern CTK"""
    
    def __init__(self, root, on_giris_basarili):
        self.root = root
        self.on_giris_basarili = on_giris_basarili
        self.parolalari_yukle()
        
        self.root.title("ŞAİK Kütüphane")
        self.root.geometry("400x450")
        self.root.resizable(False, False)
        
        if CTK_DESTEGI:
            self.root.configure(fg_color="#1a1a2e")
            self._build_ctk_ui()
        else:
            self.root.configure(bg="#1a1a2e")
            self._build_tk_ui()
    
    def _build_ctk_ui(self):
        """CustomTkinter modern UI"""
        main = ctk.CTkFrame(self.root, fg_color="#1a1a2e")
        main.pack(expand=True, fill="both", padx=40, pady=40)
        
        ctk.CTkLabel(main, text="📚", font=("Arial", 52), text_color="white").pack(pady=(20, 10))
        ctk.CTkLabel(main, text="ŞAİK KÜTÜPHANE", font=("Arial", 22, "bold"), text_color="#3b82f6").pack()
        ctk.CTkLabel(main, text="Yönetim Sistemi", font=("Arial", 12), text_color="#888").pack(pady=(0, 30))
        
        ctk.CTkButton(main, text="👨‍🏫  ÖĞRETMEN GİRİŞİ", font=("Arial", 14, "bold"),
                      width=250, height=50, corner_radius=10,
                      fg_color="#2563eb", hover_color="#1d4ed8",
                      command=lambda: self.parola_sor("ogretmen")).pack(pady=10)
        
        ctk.CTkButton(main, text="👨‍🎓  ÖĞRENCİ GİRİŞİ", font=("Arial", 14, "bold"),
                      width=250, height=50, corner_radius=10,
                      fg_color="#0891b2", hover_color="#0e7490",
                      command=lambda: self.parola_sor("ogrenci")).pack(pady=10)
        
        ctk.CTkLabel(main, text="© 2026 ŞAİK", font=("Arial", 10), text_color="#555").pack(side="bottom", pady=10)
    
    def _build_tk_ui(self):
        """Fallback Tkinter UI"""
        main = tk.Frame(self.root, bg="#1a1a2e")
        main.pack(expand=True, fill="both", padx=30, pady=30)
        tk.Label(main, text="📚", font=("Arial", 42), bg="#1a1a2e", fg="white").pack(pady=(10, 5))
        tk.Label(main, text="ŞAİK KÜTÜPHANE", font=("Arial", 20, "bold"), bg="#1a1a2e", fg="#3b82f6").pack()
        tk.Button(main, text="ÖĞRETMEN GİRİŞİ", font=("Arial", 12, "bold"), bg="#2563eb", fg="white",
                  width=22, height=2, command=lambda: self.parola_sor("ogretmen")).pack(pady=10)
        tk.Button(main, text="ÖĞRENCİ GİRİŞİ", font=("Arial", 12, "bold"), bg="#0891b2", fg="white",
                  width=22, height=2, command=lambda: self.parola_sor("ogrenci")).pack(pady=10)
    
    def parolalari_yukle(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, "okul_kutuphanesi_pro_v7.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("""CREATE TABLE IF NOT EXISTS parolalar (tip TEXT PRIMARY KEY, parola TEXT NOT NULL)""")
        cursor.execute("INSERT OR IGNORE INTO parolalar VALUES ('ogretmen', 'saik2026')")
        cursor.execute("INSERT OR IGNORE INTO parolalar VALUES ('ogrenci', 'ogrenci+')")
        conn.commit()
        cursor.execute("SELECT parola FROM parolalar WHERE tip='ogretmen'")
        self.OGRETMEN_PAROLA = cursor.fetchone()[0]
        cursor.execute("SELECT parola FROM parolalar WHERE tip='ogrenci'")
        self.OGRENCI_PAROLA = cursor.fetchone()[0]
        conn.close()
    
    def parola_sor(self, kullanici_tipi):
        if CTK_DESTEGI:
            dialog = ctk.CTkToplevel(self.root)
            dialog.title("Parola")
            dialog.geometry("320x200")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            
            baslik = "Öğretmen Parolası" if kullanici_tipi == "ogretmen" else "Öğrenci Parolası"
            ctk.CTkLabel(dialog, text=baslik, font=("Arial", 14, "bold")).pack(pady=25)
            
            parola_entry = ctk.CTkEntry(dialog, show="*", font=("Arial", 14), width=200, justify="center")
            parola_entry.pack(pady=10)
            parola_entry.focus_set()
            
            def giris_yap(event=None):
                if parola_entry.get() == (self.OGRETMEN_PAROLA if kullanici_tipi == "ogretmen" else self.OGRENCI_PAROLA):
                    dialog.destroy()
                    self.on_giris_basarili(kullanici_tipi)
                else:
                    messagebox.showerror("Hata", "Yanlış parola!", parent=dialog)
                    parola_entry.delete(0, "end")
            
            parola_entry.bind("<Return>", giris_yap)
            ctk.CTkButton(dialog, text="GİRİŞ", font=("Arial", 12, "bold"), width=120,
                          fg_color="#10b981", hover_color="#059669", command=giris_yap).pack(pady=15)
        else:
            # Fallback
            parola_pencere = tk.Toplevel(self.root)
            parola_pencere.title("Parola")
            parola_pencere.geometry("300x180")
            parola_pencere.configure(bg="#16213e")
            baslik = "Öğretmen Parolası" if kullanici_tipi == "ogretmen" else "Öğrenci Parolası"
            tk.Label(parola_pencere, text=baslik, font=("Arial", 12, "bold"), bg="#16213e", fg="white").pack(pady=20)
            parola_entry = tk.Entry(parola_pencere, show="*", font=("Arial", 12), width=20, justify="center")
            parola_entry.pack(pady=5)
            parola_entry.focus_set()
            def giris_yap(event=None):
                if parola_entry.get() == (self.OGRETMEN_PAROLA if kullanici_tipi == "ogretmen" else self.OGRENCI_PAROLA):
                    parola_pencere.destroy()
                    self.on_giris_basarili(kullanici_tipi)
                else:
                    messagebox.showerror("Hata", "Yanlış parola!", parent=parola_pencere)
            parola_entry.bind("<Return>", giris_yap)
            tk.Button(parola_pencere, text="GİRİŞ", bg="#10b981", fg="white", command=giris_yap).pack(pady=15)


class KutuphaneUygulamasi:
    def __init__(self, root, kullanici_tipi="ogretmen"):
        self.root = root
        self.kullanici_tipi = kullanici_tipi  # "ogretmen" veya "ogrenci"
        
        baslik = "ŞAİK Kütüphane Yönetim Sistemi"
        if kullanici_tipi == "ogrenci":
            baslik += " (Öğrenci Modu - Salt Okunur)"
        self.root.title(baslik)
        self.root.geometry("1280x768")

        # --- İKON AYARI ---
        self.uygulama_ikonu_ayarla()

        # --- RENK PALETİ ---
        self.bg_color = "#EAEDED"     
        self.panel_color = "#2C3E50"  
        self.accent_color = "#2980B9" 
        self.action_color = "#27AE60" 
        self.danger_color = "#C0392B" 
        self.warning_color = "#F39C12" 
        
        self.root.configure(bg=self.bg_color)

        # --- STİL AYARLARI ---
        self.style = ttk.Style()
        self.style.theme_use('clam') 
        self.style.configure("TFrame", background=self.bg_color)
        self.style.configure("TLabel", background=self.panel_color, foreground="white", font=("Segoe UI", 10))
        self.style.configure("TButton", font=("Segoe UI", 10, "bold"), borderwidth=0, focuscolor="none")
        
        self.style.configure("Ekle.TButton", background=self.accent_color, foreground="white")
        self.style.map("Ekle.TButton", background=[('active', '#3498DB')])
        self.style.configure("Islem.TButton", background=self.action_color, foreground="white", font=("Segoe UI", 11, "bold"))
        self.style.map("Islem.TButton", background=[('active', '#2ECC71')])
        self.style.configure("Sil.TButton", background=self.danger_color, foreground="white")
        self.style.map("Sil.TButton", background=[('active', '#E74C3C')])
        self.style.configure("Rapor.TButton", background=self.warning_color, foreground="white")
        self.style.map("Rapor.TButton", background=[('active', '#F1C40F')])
        self.style.configure("Normal.TButton", background="#95A5A6", foreground="white")
        self.style.configure("Devre.TButton", background="#BDC3C7", foreground="#7F8C8D")

        # Veritabanı Başlat
        self.db_adi = "okul_kutuphanesi_pro_v7.db"
        self.veritabani_kur()

        # Üst Menü
        self.menu_olustur()

        # Arayüzü Kur
        self.arayuz_olustur()
        self.verileri_guncelle()
        
        self.context_menu = tk.Menu(self.root, tearoff=0)

    # --- İKON AYARLAMA ---
    def uygulama_ikonu_ayarla(self):
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(base_dir, "logo.png")
            if os.path.exists(logo_path):
                img = tk.PhotoImage(file=logo_path)
                self.root.iconphoto(False, img)
                try:
                    import ctypes
                    myappid = 'saik.kutuphane.yonetim.v7'
                    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
                except: pass
        except Exception as e: print(f"İkon hatası: {e}")

    def tr_upper(self, text):
        if not text: return ""
        text = str(text)
        tr_map = {'ç': 'c~', 'Ç': 'C~', 'ğ': 'g~', 'Ğ': 'G~', 'ı': 'h~', 'I': 'H~',
                  'i': 'i', 'İ': 'I~~', 'ö': 'o~', 'Ö': 'O~', 'ş': 's~', 'Ş': 'S~', 'ü': 'u~', 'Ü': 'U~'}
        for key, val in tr_map.items(): text = text.replace(key, val)
        return text.lower()

    def veritabani_kur(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, self.db_adi)
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS kitaplar (id INTEGER PRIMARY KEY AUTOINCREMENT, ad TEXT NOT NULL, yazar TEXT NOT NULL, tur TEXT, sayfa_sayisi INTEGER, raf_no TEXT, durum TEXT DEFAULT 'Mevcut', barkod TEXT UNIQUE, adet INTEGER DEFAULT 1)""")
        
        # Barkod ve adet sütunları yoksa ekle (eski veritabanları için)
        try:
            self.cursor.execute("ALTER TABLE kitaplar ADD COLUMN barkod TEXT UNIQUE")
            self.conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            self.cursor.execute("ALTER TABLE kitaplar ADD COLUMN adet INTEGER DEFAULT 1")
            self.conn.commit()
        except sqlite3.OperationalError:
            pass
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS odunc_alanlar (id INTEGER PRIMARY KEY AUTOINCREMENT, kitap_id INTEGER, ogrenci_ad TEXT, ogrenci_no TEXT, sinif TEXT, alinma_tarihi TEXT, iade_tarihi TEXT, FOREIGN KEY(kitap_id) REFERENCES kitaplar(id))""")
        
        # --- YENİ: GEÇMİŞ TABLOSU (WRAPPED İÇİN) ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS odunc_gecmisi (id INTEGER PRIMARY KEY AUTOINCREMENT, kitap_ad TEXT, yazar TEXT, tur TEXT, ogrenci_ad TEXT, sinif TEXT, alinma_tarihi TEXT, iade_tarihi TEXT)""")
        
        # --- YENİ: ÖĞRENCİLER TABLOSU ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS ogrenciler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            okul_no TEXT UNIQUE NOT NULL,
            ad_soyad TEXT NOT NULL,
            sinif TEXT
        )""")
        
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS ayarlar (anahtar TEXT PRIMARY KEY, deger TEXT)""")
        self.cursor.execute("INSERT OR IGNORE INTO ayarlar (anahtar, deger) VALUES ('odunc_suresi', '45')")
        self.conn.commit()

    def menu_olustur(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        dosya_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Dosya", menu=dosya_menu)
        
        # Excel Import (herkes için)
        if EXCEL_DESTEGI:
            dosya_menu.add_command(label="📥 Excel'den Kitap Aktar", command=self.excel_import)
        else:
            dosya_menu.add_command(label="📥 Excel'den Kitap Aktar (openpyxl yükleyin)", state="disabled")
        dosya_menu.add_separator()
        
        dosya_menu.add_command(label="💾 Veritabanını Yedekle", command=self.yedekle)
        dosya_menu.add_command(label="🏆 Yıllık Özet", command=self.wrapped_penceresi)
        dosya_menu.add_separator()
        dosya_menu.add_command(label="🚪 Çıkış Yap", command=self.cikis_yap)
        
        # Ayarlar Menüsü (sadece öğretmen)
        if self.kullanici_tipi == "ogretmen":
            ayar_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="⚙️ Ayarlar", menu=ayar_menu)
            ayar_menu.add_command(label="� Parola Değiştir", command=self.parola_degistir_penceresi)
            ayar_menu.add_command(label="📅 Ödünç Süresi", command=self.ayarlar_penceresi)
        
        # Barkod Menüsü
        barkod_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📊 Barkod İşlemleri", menu=barkod_menu)
        
        if self.kullanici_tipi == "ogretmen":
            if BARKOD_OLUSTURMA_DESTEGI:
                barkod_menu.add_command(label="🏷️ Seçili Kitap İçin Barkod Oluştur", command=self.barkod_olustur)
                barkod_menu.add_command(label="🏷️ Tüm Kitaplara Barkod Oluştur", command=self.toplu_barkod_olustur)
            else:
                barkod_menu.add_command(label="🏷️ Barkod Oluştur (python-barcode yükleyin)", state="disabled")
            barkod_menu.add_separator()
        
        if BARKOD_OKUMA_DESTEGI:
            barkod_menu.add_command(label="📷 Barkod Tara (Kamera)", command=self.barkod_tara)
        else:
            barkod_menu.add_command(label="📷 Barkod Tara (opencv-python & pyzbar yükleyin)", state="disabled")
        
        barkod_menu.add_separator()
        barkod_menu.add_command(label="🔍 Barkod ile Ara", command=self.barkod_ile_ara)
        
        # Öğrenciler Menüsü
        ogrenci_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="👥 Öğrenciler", menu=ogrenci_menu)
        
        if EXCEL_DESTEGI:
            ogrenci_menu.add_command(label="📥 Excel'den Öğrenci Aktar", command=self.ogrenci_excel_import)
        else:
            ogrenci_menu.add_command(label="📥 Excel'den Öğrenci Aktar (openpyxl yükleyin)", state="disabled")
        
        ogrenci_menu.add_separator()
        ogrenci_menu.add_command(label="👥 Öğrenci Listesi", command=self.ogrenci_listesi_penceresi)

    def arayuz_olustur(self):
        # SOL PANEL
        left_panel = tk.Frame(self.root, bg=self.panel_color, width=320)
        left_panel.pack(side=tk.LEFT, fill=tk.Y)
        left_panel.pack_propagate(False)

        header_frame = tk.Frame(left_panel, bg=self.panel_color)
        header_frame.pack(pady=25)

        tk.Label(header_frame, text="ŞAİK", bg=self.panel_color, fg="white", font=("Segoe UI", 28, "bold")).pack()
        tk.Label(header_frame, text="KÜTÜPHANE", bg=self.panel_color, fg="#BDC3C7", font=("Segoe UI", 14, "bold")).pack()

        stats_frame = tk.Frame(left_panel, bg="#34495E", pady=10)
        stats_frame.pack(fill=tk.X, padx=15, pady=15)
        self.lbl_toplam_kitap = tk.Label(stats_frame, text="Toplam: 0", bg="#34495E", fg="white", font=("Segoe UI", 9, "bold"))
        self.lbl_toplam_kitap.pack(anchor="w", padx=10)
        self.lbl_odunc_kitap = tk.Label(stats_frame, text="Ödünçte: 0", bg="#34495E", fg="#F1C40F", font=("Segoe UI", 9, "bold"))
        self.lbl_odunc_kitap.pack(anchor="w", padx=10)
        self.lbl_gecikmis_kitap = tk.Label(stats_frame, text="Gecikmiş: 0", bg="#34495E", fg="#E74C3C", font=("Segoe UI", 9, "bold"))
        self.lbl_gecikmis_kitap.pack(anchor="w", padx=10)

        form_frame = tk.Frame(left_panel, bg=self.panel_color)
        form_frame.pack(fill=tk.BOTH, expand=True, padx=20)
        self.entry_ad = self.create_input(form_frame, "Kitap Adı")
        self.entry_yazar = self.create_input(form_frame, "Yazar")
        self.entry_tur = self.create_input(form_frame, "Tür")
        self.entry_sayfa = self.create_input(form_frame, "Sayfa Sayısı")
        self.entry_raf = self.create_input(form_frame, "Raf No")
        self.entry_adet = self.create_input(form_frame, "Adet")
        self.entry_adet.insert(0, "1")  # Varsayılan 1

        btn_frame = tk.Frame(left_panel, bg=self.panel_color)
        btn_frame.pack(fill=tk.X, padx=20, pady=20, side=tk.BOTTOM)
        
        # Öğretmen: tüm butonlar aktif
        if self.kullanici_tipi == "ogretmen":
            ttk.Button(btn_frame, text="⚡ ÖDÜNÇ VER / İADE AL", command=self.akilli_islem_yap, style="Islem.TButton", cursor="hand2").pack(fill=tk.X, pady=5, ipady=3)
        
        # Herkes kitap ekleyebilir
        ttk.Button(btn_frame, text="➕ KİTAP EKLE", command=self.kitap_ekle, style="Ekle.TButton", cursor="hand2").pack(fill=tk.X, pady=5)
        
        if self.kullanici_tipi == "ogretmen":
            ttk.Button(btn_frame, text="�️ SİL", command=self.kitap_sil, style="Sil.TButton", cursor="hand2").pack(fill=tk.X, pady=5)
        
        ttk.Button(btn_frame, text="� TEMİZLE", command=self.formu_temizle, style="Normal.TButton", cursor="hand2").pack(fill=tk.X, pady=5)

        # SAĞ PANEL
        right_panel = tk.Frame(self.root, bg=self.bg_color)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        top_bar = tk.Frame(right_panel, bg="#FFFFFF", padx=10, pady=10)
        top_bar.pack(fill=tk.X)
        tk.Label(top_bar, text="🔍 Hızlı Ara:", bg="#FFFFFF", fg="#7F8C8D", font=("Segoe UI", 11)).pack(side=tk.LEFT)
        self.entry_ara = ttk.Entry(top_bar, font=("Segoe UI", 11))
        self.entry_ara.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        self.entry_ara.bind("<KeyRelease>", self.arama_yap)
        tk.Label(top_bar, text="Sıralama:", bg="#FFFFFF", fg="#7F8C8D", font=("Segoe UI", 11)).pack(side=tk.LEFT, padx=(10, 5))
        self.sort_var = tk.StringVar(value="Ekleme Sırası (Yeniden Eskiye)")
        self.combo_sort = ttk.Combobox(top_bar, textvariable=self.sort_var, state="readonly", width=25, font=("Segoe UI", 10))
        self.combo_sort['values'] = ("Ekleme Sırası (Yeniden Eskiye)", "Ekleme Sırası (Eskiden Yeniye)", "Kitap Adı (A-Z)", "Yazar Adı (A-Z)")
        self.combo_sort.pack(side=tk.LEFT, padx=5)
        self.combo_sort.bind("<<ComboboxSelected>>", lambda e: self.verileri_guncelle())
        ttk.Button(top_bar, text="Tümü", command=lambda: self.verileri_guncelle(), style="Normal.TButton").pack(side=tk.RIGHT, padx=2)
        ttk.Button(top_bar, text="Ödünçtekiler", command=lambda: self.filtrele("Ödünç"), style="Normal.TButton").pack(side=tk.RIGHT, padx=2)

        tree_frame = tk.Frame(right_panel, bg=self.bg_color, padx=10, pady=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        columns = ("ID", "Ad", "Yazar", "Tur", "Sayfa", "Raf", "Durum", "IadeTarihi")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        headers = ["ID", "Kitap Adı", "Yazar", "Tür", "Sayfa", "Raf", "Durum", "Son İade Tarihi"]
        widths = [40, 250, 150, 100, 70, 70, 100, 100]
        for col, h, w in zip(columns, headers, widths):
            self.tree.heading(col, text=h)
            self.tree.column(col, width=w, anchor="center" if col != "Ad" else "w")
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.tag_configure('odd', background='#F7F9F9') 
        self.tree.tag_configure('even', background='#FFFFFF')
        self.tree.tag_configure('odunc_normal', background='#FDEBD0', foreground='#D35400')
        self.tree.tag_configure('odunc_yaklasan', background='#FFF9C4', foreground='#FBC02D')
        self.tree.tag_configure('odunc_gecikmis', background='#FADBD8', foreground='#C0392B')
        self.tree.bind("<Button-3>", self.sag_tik_goster)
        self.tree.bind("<Double-1>", lambda e: self.akilli_islem_yap()) 
        self.status_bar = tk.Label(self.root, text="Sistem Hazır", bg="#ECF0F1", fg="#7F8C8D", anchor="w", padx=10, font=("Segoe UI", 9))
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def create_input(self, parent, title):
        tk.Label(parent, text=title, anchor="w").pack(fill=tk.X, pady=(10,2))
        entry = ttk.Entry(parent, font=("Segoe UI", 10))
        entry.pack(fill=tk.X, ipady=3) 
        return entry

    def durum_yaz(self, mesaj):
        self.status_bar.config(text=f"Bilgi: {mesaj}")
        self.root.after(3000, lambda: self.status_bar.config(text="Sistem Hazır"))

    # --- WRAPPED ÖZELLİĞİ ---
    def wrapped_penceresi(self):
        top = tk.Toplevel(self.root)
        top.title("ŞAİK Wrapped - Yıllık Özet")
        top.geometry("800x600")
        top.configure(bg="#2C3E50")

        # Başlık ve Yıl Seçimi
        header = tk.Frame(top, bg="#2C3E50")
        header.pack(pady=20)
        tk.Label(header, text="🏆 YILLIK KÜTÜPHANE ÖZETİ", font=("Segoe UI", 24, "bold"), bg="#2C3E50", fg="#F1C40F").pack()
        
        yil_frame = tk.Frame(top, bg="#2C3E50")
        yil_frame.pack(pady=10)
        tk.Label(yil_frame, text="Yıl Seçiniz:", fg="white", bg="#2C3E50", font=("Segoe UI", 12)).pack(side=tk.LEFT, padx=10)
        
        current_year = str(datetime.now().year)
        yil_combo = ttk.Combobox(yil_frame, values=[str(y) for y in range(2024, 2030)], width=10, state="readonly")
        yil_combo.set(current_year)
        yil_combo.pack(side=tk.LEFT)

        content_frame = tk.Frame(top, bg="#2C3E50")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        def istatistikleri_getir():
            yil = yil_combo.get()
            # Öncekileri temizle
            for widget in content_frame.winfo_children(): widget.destroy()

            # Verileri Çek (Hem aktif ödünçler hem geçmiş)
            veriler = []
            
            # 1. Aktif Ödünçlerden
            self.cursor.execute("SELECT k.ad, k.yazar, k.tur, o.ogrenci_ad, o.alinma_tarihi FROM odunc_alanlar o JOIN kitaplar k ON o.kitap_id = k.id")
            for row in self.cursor.fetchall():
                if row[4] and row[4].endswith(yil): veriler.append(row)
            
            # 2. Geçmişten
            self.cursor.execute("SELECT kitap_ad, yazar, tur, ogrenci_ad, alinma_tarihi FROM odunc_gecmisi")
            for row in self.cursor.fetchall():
                if row[4] and row[4].endswith(yil): veriler.append(row)

            if not veriler:
                tk.Label(content_frame, text=f"{yil} yılına ait veri bulunamadı.", bg="#2C3E50", fg="white", font=("Segoe UI", 14)).pack(pady=50)
                return

            # Hesaplamalar
            kitaplar = [v[0] for v in veriler]
            yazarlar = [v[1] for v in veriler]
            turler = [v[2] for v in veriler]
            ogrenciler = [v[3] for v in veriler]

            top_kitap = Counter(kitaplar).most_common(1)[0]
            top_yazar = Counter(yazarlar).most_common(1)[0]
            top_tur = Counter(turler).most_common(1)[0]
            top_ogrenci = Counter(ogrenciler).most_common(1)[0]
            toplam_okunan = len(veriler)

            # Kartları Oluştur
            self.kart_olustur(content_frame, "📚 YILIN KİTABI", f"{top_kitap[0]}\n({top_kitap[1]} kez okundu)", "#E74C3C", 0, 0)
            self.kart_olustur(content_frame, "✍️ YILIN YAZARI", f"{top_yazar[0]}\n({top_yazar[1]} kitap)", "#8E44AD", 0, 1)
            self.kart_olustur(content_frame, "🎭 EN SEVİLEN TÜR", f"{top_tur[0]}", "#2980B9", 1, 0)
            self.kart_olustur(content_frame, "🎓 KİTAP KURDU", f"{top_ogrenci[0]}\n({top_ogrenci[1]} kitap okudu)", "#F1C40F", 1, 1)
            
            tk.Label(content_frame, text=f"Bu yıl toplam {toplam_okunan} kitap ödünç verildi!", font=("Segoe UI", 12, "italic"), bg="#2C3E50", fg="#BDC3C7").grid(row=2, column=0, columnspan=2, pady=20)

        ttk.Button(yil_frame, text="GÖSTER", command=istatistikleri_getir, style="Islem.TButton").pack(side=tk.LEFT, padx=10)
        
        # İlk açılışta verileri getir
        istatistikleri_getir()

    def kart_olustur(self, parent, baslik, icerik, renk, r, c):
        frame = tk.Frame(parent, bg=renk, padx=5, pady=5)
        frame.grid(row=r, column=c, padx=10, pady=10, sticky="nsew")
        parent.grid_columnconfigure(c, weight=1)
        parent.grid_rowconfigure(r, weight=1)
        
        tk.Label(frame, text=baslik, bg=renk, fg="white", font=("Segoe UI", 12, "bold")).pack(pady=(10, 5))
        tk.Label(frame, text=icerik, bg=renk, fg="white", font=("Segoe UI", 14), wraplength=300).pack(pady=10)

    # --- YEDEKLEME ---
    def yedekle(self):
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            src_file = os.path.join(base_dir, self.db_adi)
            hedef_dosya = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("Veritabanı Dosyası", "*.db")], initialfile=f"Yedek_{datetime.now().strftime('%Y%m%d')}_{self.db_adi}", title="Yedeği Kaydet")
            if hedef_dosya:
                shutil.copy2(src_file, hedef_dosya)
                messagebox.showinfo("Başarılı", f"Yedekleme tamamlandı:\n{hedef_dosya}")
        except Exception as e: messagebox.showerror("Hata", f"Yedekleme hatası: {e}")

    def ayarlar_penceresi(self):
        top = tk.Toplevel(self.root)
        top.title("Sistem Ayarları")
        top.geometry("300x200")
        top.configure(bg=self.bg_color)
        self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar='odunc_suresi'")
        mevcut_sure = self.cursor.fetchone()[0]
        tk.Label(top, text="Ödünç Verme Süresi (Gün)", bg=self.bg_color, font=("Segoe UI", 10, "bold")).pack(pady=20)
        entry_sure = ttk.Entry(top, font=("Segoe UI", 12), justify='center')
        entry_sure.insert(0, mevcut_sure)
        entry_sure.pack(pady=5, padx=20)
        def kaydet():
            yeni_sure = entry_sure.get()
            if not yeni_sure.isdigit():
                messagebox.showerror("Hata", "Lütfen geçerli bir sayı girin.", parent=top)
                return
            self.cursor.execute("UPDATE ayarlar SET deger=? WHERE anahtar='odunc_suresi'", (yeni_sure,))
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Ayarlar kaydedildi. Yeni işlemler bu süreye göre yapılacak.", parent=top)
            top.destroy()
        ttk.Button(top, text="KAYDET", command=kaydet, style="Islem.TButton").pack(pady=20, fill=tk.X, padx=20)

    def verileri_guncelle(self):
        self.tree.delete(*self.tree.get_children())
        query = """SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id"""
        self.cursor.execute(query)
        rows = self.cursor.fetchall()
        sort_option = self.combo_sort.get()
        if sort_option == "Ekleme Sırası (Yeniden Eskiye)": rows.sort(key=lambda x: x[0], reverse=True) 
        elif sort_option == "Ekleme Sırası (Eskiden Yeniye)": rows.sort(key=lambda x: x[0], reverse=False) 
        elif sort_option == "Kitap Adı (A-Z)": rows.sort(key=lambda x: self.tr_upper(x[1])) 
        elif sort_option == "Yazar Adı (A-Z)": rows.sort(key=lambda x: self.tr_upper(x[2])) 
        odunc_sayisi = 0
        gecikmis_sayisi = 0
        bugun = datetime.now()
        for i, row in enumerate(rows):
            durum = row[6]
            iade_tarihi_str = row[7]
            tag = 'even' if i % 2 == 0 else 'odd'
            if durum != 'Mevcut':
                odunc_sayisi += 1
                if iade_tarihi_str:
                    try:
                        iade_tarihi = datetime.strptime(iade_tarihi_str, "%d.%m.%Y")
                        kalan_gun = (iade_tarihi - bugun).days
                        if kalan_gun < 0:
                            tag = 'odunc_gecikmis'
                            gecikmis_sayisi += 1
                        elif kalan_gun <= 3: tag = 'odunc_yaklasan'
                        else: tag = 'odunc_normal'
                    except: tag = 'odunc_normal'
                else: tag = 'odunc_normal'
            display_row = list(row)
            if display_row[7] is None: display_row[7] = "-"
            self.tree.insert("", tk.END, values=display_row, tags=(tag,))
        self.lbl_toplam_kitap.config(text=f"Toplam Kitap: {len(rows)}")
        self.lbl_odunc_kitap.config(text=f"Ödünçte: {odunc_sayisi}")
        self.lbl_gecikmis_kitap.config(text=f"Gecikmiş: {gecikmis_sayisi}")

    def akilli_islem_yap(self):
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen listeden bir kitap seçiniz.")
            return
        item = self.tree.item(secili)
        durum = item['values'][6]
        if durum == 'Mevcut': self.odunc_ver_penceresi()
        else:
            popup = tk.Toplevel(self.root)
            popup.title("İşlem Seç")
            popup.geometry("300x150")
            popup.configure(bg=self.bg_color)
            tk.Label(popup, text=f"Seçili Kitap: {item['values'][1]}", bg=self.bg_color, font=("Segoe UI", 10, "bold")).pack(pady=10)
            ttk.Button(popup, text="ℹ️ KİMDE? (Bilgi Göster)", command=lambda: [self.odunc_bilgisi_goster(), popup.destroy()], style="Normal.TButton").pack(fill=tk.X, padx=20, pady=5)
            ttk.Button(popup, text="✅ İADE AL (Rafa Kaldır)", command=lambda: [self.iade_al(), popup.destroy()], style="Islem.TButton").pack(fill=tk.X, padx=20, pady=5)

    def kitap_ekle(self):
        adet_str = self.entry_adet.get().strip()
        adet = int(adet_str) if adet_str.isdigit() and int(adet_str) > 0 else 1
        
        veriler = (self.entry_ad.get(), self.entry_yazar.get(), self.entry_tur.get(), 
                   self.entry_sayfa.get(), self.entry_raf.get(), adet)
        if not veriler[0] or not veriler[1]:
            messagebox.showwarning("Eksik Bilgi", "Kitap Adı ve Yazar alanları zorunludur.")
            return
        try:
            self.cursor.execute("INSERT INTO kitaplar (ad, yazar, tur, sayfa_sayisi, raf_no, adet) VALUES (?,?,?,?,?,?)", veriler)
            self.conn.commit()
            self.verileri_guncelle()
            self.formu_temizle()
            self.durum_yaz(f"Kitap eklendi ({adet} adet).")
        except Exception as e: messagebox.showerror("Hata", str(e))

    def kitap_sil(self):
        secili = self.tree.selection()
        if secili:
            item = self.tree.item(secili)
            if messagebox.askyesno("Sil", f"'{item['values'][1]}' kitabını silmek istediğinize emin misiniz?"):
                id = item['values'][0]
                self.cursor.execute("DELETE FROM kitaplar WHERE id=?", (id,))
                self.cursor.execute("DELETE FROM odunc_alanlar WHERE kitap_id=?", (id,))
                self.conn.commit()
                self.verileri_guncelle()
                self.durum_yaz("Kitap silindi.")

    def odunc_ver_penceresi(self):
        secili = self.tree.selection()
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        top = tk.Toplevel(self.root)
        top.title("Ödünç Verme İşlemi")
        top.geometry("400x400")
        top.configure(bg=self.bg_color)
        tk.Label(top, text="Öğrenci Bilgileri", font=("Segoe UI", 14, "bold"), bg=self.bg_color, fg=self.panel_color).pack(pady=20)
        entries = {}
        for alan in ["Öğrenci Adı Soyadı", "Okul No", "Sınıf"]:
            frame = tk.Frame(top, bg=self.bg_color)
            frame.pack(fill=tk.X, padx=30, pady=5)
            tk.Label(frame, text=alan, bg=self.bg_color, width=15, anchor="w").pack(side=tk.LEFT)
            e = ttk.Entry(frame)
            e.pack(side=tk.RIGHT, fill=tk.X, expand=True)
            entries[alan] = e
        def onayla():
            if not all(e.get() for e in entries.values()):
                messagebox.showwarning("Eksik", "Tüm alanları doldurunuz.", parent=top)
                return
            self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar='odunc_suresi'")
            gun_sayisi = int(self.cursor.fetchone()[0])
            bugun = datetime.now()
            iade = bugun + timedelta(days=gun_sayisi)
            tarih_fmt = "%d.%m.%Y"
            self.cursor.execute("UPDATE kitaplar SET durum='Ödünç Verildi' WHERE id=?", (kitap_id,))
            self.cursor.execute("INSERT INTO odunc_alanlar (kitap_id, ogrenci_ad, ogrenci_no, sinif, alinma_tarihi, iade_tarihi) VALUES (?,?,?,?,?,?)", 
                                (kitap_id, entries["Öğrenci Adı Soyadı"].get(), entries["Okul No"].get(), entries["Sınıf"].get(), bugun.strftime(tarih_fmt), iade.strftime(tarih_fmt)))
            self.conn.commit()
            self.verileri_guncelle()
            top.destroy()
            self.durum_yaz(f"Kitap verildi. Son iade tarihi: {iade.strftime(tarih_fmt)}")
            messagebox.showinfo("Başarılı", f"İşlem Tamam!\nÖğrenciye '{iade.strftime(tarih_fmt)}' tarihine kadar süre verildi ({gun_sayisi} Gün).")
        ttk.Button(top, text="ONAYLA VE VER", command=onayla, style="Ekle.TButton").pack(fill=tk.X, padx=30, pady=30)

    def iade_al(self):
        secili = self.tree.selection()
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        if messagebox.askyesno("İade Onayı", f"'{item['values'][1]}' kitabı fiziksel olarak teslim alındı mı?"):
            # 1. Önce verileri al
            self.cursor.execute("""
                SELECT k.ad, k.yazar, k.tur, o.ogrenci_ad, o.sinif, o.alinma_tarihi 
                FROM odunc_alanlar o 
                JOIN kitaplar k ON o.kitap_id = k.id 
                WHERE o.kitap_id = ?""", (kitap_id,))
            veri = self.cursor.fetchone()
            
            # 2. Arşive (Geçmişe) Kaydet (Eğer veri varsa)
            if veri:
                bugun = datetime.now().strftime("%d.%m.%Y")
                self.cursor.execute("""
                    INSERT INTO odunc_gecmisi (kitap_ad, yazar, tur, ogrenci_ad, sinif, alinma_tarihi, iade_tarihi)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (veri[0], veri[1], veri[2], veri[3], veri[4], veri[5], bugun))

            # 3. İade işlemini yap (Sil ve durumu güncelle)
            self.cursor.execute("UPDATE kitaplar SET durum='Mevcut' WHERE id=?", (kitap_id,))
            self.cursor.execute("DELETE FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
            self.conn.commit()
            self.verileri_guncelle()
            self.durum_yaz("Kitap iade alındı ve geçmişe kaydedildi.")

    def rapor_al(self):
        try:
            dosya_yolu = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Dosyası", "*.csv"), ("Tüm Dosyalar", "*.*")], title="Raporu Kaydet")
            if not dosya_yolu: return
            query = """SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.ogrenci_ad, o.ogrenci_no, o.sinif, o.alinma_tarihi, o.iade_tarihi 
                       FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id ORDER BY k.id ASC"""
            self.cursor.execute(query)
            rows = self.cursor.fetchall()
            with open(dosya_yolu, mode='w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file, delimiter=';')
                writer.writerow(["ID", "Kitap Adı", "Yazar", "Tür", "Sayfa", "Raf No", "Durum", "Öğrenci Adı", "Öğrenci No", "Sınıf", "Veriliş Tarihi", "Son İade Tarihi"])
                for row in rows: writer.writerow(row)
            messagebox.showinfo("Başarılı", f"Rapor kaydedildi:\n{dosya_yolu}")
        except Exception as e: messagebox.showerror("Hata", f"Rapor hatası: {e}")

    def odunc_bilgisi_goster(self):
        secili = self.tree.selection()
        kitap_id = self.tree.item(secili)['values'][0]
        self.cursor.execute("SELECT * FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
        bilgi = self.cursor.fetchone()
        if bilgi: messagebox.showinfo("Teslim Bilgisi", f"Öğrenci: {bilgi[2]}\nNo: {bilgi[3]}\nSınıf: {bilgi[4]}\n\nVeriliş: {bilgi[5]}\nSon Tarih: {bilgi[6]}")

    def sag_tik_goster(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            durum = self.tree.item(item)['values'][6]
            self.context_menu.delete(0, tk.END)
            if durum == 'Mevcut':
                self.context_menu.add_command(label="➕ Ödünç Ver", command=self.odunc_ver_penceresi)
            else:
                self.context_menu.add_command(label="ℹ️ Kimde? (Bilgi)", command=self.odunc_bilgisi_goster)
                self.context_menu.add_command(label="✅ İade Al", command=self.iade_al)
            self.context_menu.add_separator()
            self.context_menu.add_command(label="🗑️ Sil", command=self.kitap_sil)
            self.context_menu.post(event.x_root, event.y_root)

    def formu_temizle(self):
        for e in [self.entry_ad, self.entry_yazar, self.entry_tur, self.entry_sayfa, self.entry_raf, self.entry_adet]: 
            e.delete(0, tk.END)
        self.entry_adet.insert(0, "1")
        self.durum_yaz("Form temizlendi.")

    def filtrele(self, mod):
        self.tree.delete(*self.tree.get_children())
        if mod == "Ödünç": query = "SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id WHERE k.durum != 'Mevcut' ORDER BY k.id DESC"
        else: query = "SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id ORDER BY k.id DESC"
        self.cursor.execute(query)
        rows = self.cursor.fetchall()
        bugun = datetime.now()
        for i, row in enumerate(rows):
            durum = row[6]
            iade_tarihi_str = row[7]
            tag = 'odunc_normal'
            if durum != 'Mevcut' and iade_tarihi_str:
                 try:
                    iade_tarihi = datetime.strptime(iade_tarihi_str, "%d.%m.%Y")
                    kalan_gun = (iade_tarihi - bugun).days
                    if kalan_gun < 0: tag = 'odunc_gecikmis'
                    elif kalan_gun <= 3: tag = 'odunc_yaklasan'
                 except: pass
            display_row = list(row)
            if display_row[7] is None: display_row[7] = "-"
            self.tree.insert("", tk.END, values=display_row, tags=(tag,))

    def arama_yap(self, event):
        anahtar = self.entry_ara.get()
        self.tree.delete(*self.tree.get_children())
        query = "SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id WHERE k.ad LIKE ? OR k.yazar LIKE ?"
        self.cursor.execute(query, (f"%{anahtar}%", f"%{anahtar}%"))
        rows = self.cursor.fetchall()
        for row in rows:
            tag = 'even'
            if row[6] != 'Mevcut': tag = 'odunc_normal'
            display_row = list(row)
            if display_row[7] is None: display_row[7] = "-"
            self.tree.insert("", tk.END, values=display_row, tags=(tag,))

    # --- EXCEL IMPORT ---
    def excel_import(self):
        """Excel dosyasından kitap aktarma"""
        if not EXCEL_DESTEGI:
            messagebox.showerror("Hata", "Excel desteği için 'openpyxl' kütüphanesini yükleyin:\npip install openpyxl")
            return
        
        dosya = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )
        
        if not dosya:
            return
        
        try:
            wb = load_workbook(dosya)
            ws = wb.active
            
            # Başlıkları bul
            basliklar = []
            for cell in ws[1]:
                basliklar.append(str(cell.value).lower().strip() if cell.value else "")
            
            # Sütun eşleştirme
            sutun_map = {
                'ad': None, 'yazar': None, 'tur': None, 
                'sayfa': None, 'raf': None, 'adet': None
            }
            
            for i, baslik in enumerate(basliklar):
                if 'kitap' in baslik and 'ad' in baslik:
                    sutun_map['ad'] = i
                elif 'ad' in baslik and sutun_map['ad'] is None:
                    sutun_map['ad'] = i
                elif 'yazar' in baslik:
                    sutun_map['yazar'] = i
                elif 'tür' in baslik or 'tur' in baslik:
                    sutun_map['tur'] = i
                elif 'sayfa' in baslik:
                    sutun_map['sayfa'] = i
                elif 'raf' in baslik:
                    sutun_map['raf'] = i
                elif 'adet' in baslik or 'miktar' in baslik or 'sayı' in baslik:
                    sutun_map['adet'] = i
            
            # Verileri oku
            kitaplar = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                # Adet değerini al
                adet_val = 1
                if sutun_map['adet'] is not None and row[sutun_map['adet']]:
                    try:
                        adet_val = int(row[sutun_map['adet']])
                        if adet_val < 1: adet_val = 1
                    except: adet_val = 1
                
                kitap = {
                    'ad': str(row[sutun_map['ad']]) if sutun_map['ad'] is not None and row[sutun_map['ad']] else "",
                    'yazar': str(row[sutun_map['yazar']]) if sutun_map['yazar'] is not None and row[sutun_map['yazar']] else "",
                    'tur': str(row[sutun_map['tur']]) if sutun_map['tur'] is not None and row[sutun_map['tur']] else "",
                    'sayfa': row[sutun_map['sayfa']] if sutun_map['sayfa'] is not None else None,
                    'raf': str(row[sutun_map['raf']]) if sutun_map['raf'] is not None and row[sutun_map['raf']] else "",
                    'adet': adet_val
                }
                
                if kitap['ad']:
                    kitaplar.append(kitap)
            
            if not kitaplar:
                messagebox.showwarning("Uyarı", "Excel dosyasında geçerli kitap bulunamadı.")
                return
            
            # Onay iste
            onay = messagebox.askyesno(
                "Onay", 
                f"{len(kitaplar)} kitap bulundu.\n\nBu kitapları veritabanına eklemek istiyor musunuz?"
            )
            
            if onay:
                eklenen = 0
                for kitap in kitaplar:
                    try:
                        self.cursor.execute(
                            "INSERT INTO kitaplar (ad, yazar, tur, sayfa_sayisi, raf_no, adet) VALUES (?,?,?,?,?,?)",
                            (kitap['ad'], kitap['yazar'], kitap['tur'], kitap['sayfa'], kitap['raf'], kitap['adet'])
                        )
                        eklenen += 1
                    except Exception as e:
                        print(f"Kitap eklenemedi: {kitap['ad']} - {e}")
                
                self.conn.commit()
                self.verileri_guncelle()
                messagebox.showinfo("Başarılı", f"{eklenen} kitap başarıyla eklendi.")
                self.durum_yaz(f"Excel'den {eklenen} kitap aktarıldı.")
                
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası okunamadı:\n{e}")
    
    # --- BARKOD OLUŞTURMA ---
    def barkod_olustur(self):
        """Seçili kitap için barkod oluştur"""
        if not BARKOD_OLUSTURMA_DESTEGI:
            messagebox.showerror("Hata", "Barkod desteği için 'python-barcode' ve 'pillow' yükleyin:\npip install python-barcode pillow")
            return
        
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen listeden bir kitap seçiniz.")
            return
        
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        kitap_ad = item['values'][1]
        
        # Mevcut barkodu kontrol et
        self.cursor.execute("SELECT barkod FROM kitaplar WHERE id=?", (kitap_id,))
        mevcut = self.cursor.fetchone()[0]
        
        if mevcut:
            cevap = messagebox.askyesno("Barkod Mevcut", 
                f"Bu kitabın zaten bir barkodu var:\n{mevcut}\n\nYeni barkod oluşturmak istiyor musunuz?")
            if not cevap:
                self.barkod_goster(kitap_id)
                return
        
        # Benzersiz barkod oluştur
        barkod_kodu = self.benzersiz_barkod_olustur()
        
        # Veritabanını güncelle
        self.cursor.execute("UPDATE kitaplar SET barkod=? WHERE id=?", (barkod_kodu, kitap_id))
        self.conn.commit()
        
        # Barkod görselini kaydet
        self.barkod_kaydet(barkod_kodu, kitap_ad)
        
        messagebox.showinfo("Başarılı", f"Barkod oluşturuldu:\n{barkod_kodu}")
        self.barkod_goster(kitap_id)
    
    def benzersiz_barkod_olustur(self):
        """Benzersiz 12 haneli barkod kodu oluştur"""
        while True:
            # EAN-13 formatı için 12 hane (13. hane kontrol hanesi)
            kod = "978" + "".join([str(random.randint(0, 9)) for _ in range(9)])
            self.cursor.execute("SELECT id FROM kitaplar WHERE barkod=?", (kod,))
            if not self.cursor.fetchone():
                return kod
    
    def barkod_kaydet(self, barkod_kodu, kitap_ad):
        """Barkod görselini PNG olarak kaydet"""
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            barkod_dir = os.path.join(base_dir, "barkodlar")
            
            if not os.path.exists(barkod_dir):
                os.makedirs(barkod_dir)
            
            # Barkod oluştur
            EAN = barcode.get_barcode_class('ean13')
            ean = EAN(barkod_kodu, writer=ImageWriter())
            
            # Dosya adını temizle
            guvenli_ad = "".join(c for c in kitap_ad if c.isalnum() or c in (' ', '-', '_')).strip()[:50]
            dosya_yolu = os.path.join(barkod_dir, f"{guvenli_ad}_{barkod_kodu}")
            
            ean.save(dosya_yolu)
            return dosya_yolu + ".png"
        except Exception as e:
            print(f"Barkod kaydetme hatası: {e}")
            return None
    
    def barkod_goster(self, kitap_id):
        """Barkod görüntüleme penceresi"""
        self.cursor.execute("SELECT ad, barkod FROM kitaplar WHERE id=?", (kitap_id,))
        sonuc = self.cursor.fetchone()
        
        if not sonuc or not sonuc[1]:
            messagebox.showwarning("Uyarı", "Bu kitabın barkodu bulunamadı.")
            return
        
        kitap_ad, barkod_kodu = sonuc
        
        top = tk.Toplevel(self.root)
        top.title(f"Barkod - {kitap_ad[:30]}")
        top.geometry("400x350")
        top.configure(bg="white")
        
        tk.Label(top, text=kitap_ad, font=("Segoe UI", 12, "bold"), 
                 bg="white", wraplength=380).pack(pady=15)
        
        # Barkod görselini yükle
        base_dir = os.path.dirname(os.path.abspath(__file__))
        barkod_dir = os.path.join(base_dir, "barkodlar")
        
        barkod_dosya = None
        if os.path.exists(barkod_dir):
            for dosya in os.listdir(barkod_dir):
                if barkod_kodu in dosya:
                    barkod_dosya = os.path.join(barkod_dir, dosya)
                    break
        
        if barkod_dosya and os.path.exists(barkod_dosya) and PIL_DESTEGI:
            try:
                img = Image.open(barkod_dosya)
                img = img.resize((350, 150), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                label = tk.Label(top, image=photo, bg="white")
                label.image = photo
                label.pack(pady=10)
            except Exception as e:
                tk.Label(top, text="[Barkod görseli yüklenemedi]", bg="white", fg="gray").pack(pady=20)
        else:
            tk.Label(top, text="[Barkod görseli bulunamadı]", bg="white", fg="gray").pack(pady=20)
        
        tk.Label(top, text=f"Barkod: {barkod_kodu}", font=("Consolas", 14), bg="white").pack(pady=10)
        
        def kopyala():
            self.root.clipboard_clear()
            self.root.clipboard_append(barkod_kodu)
            messagebox.showinfo("Kopyalandı", "Barkod panoya kopyalandı.")
        
        tk.Button(top, text="📋 Kopyala", command=kopyala, 
                  font=("Segoe UI", 10), bg="#3498DB", fg="white").pack(pady=10)
    
    def toplu_barkod_olustur(self):
        """Barkodu olmayan tüm kitaplara barkod oluştur"""
        self.cursor.execute("SELECT id, ad FROM kitaplar WHERE barkod IS NULL OR barkod = ''")
        kitaplar = self.cursor.fetchall()
        
        if not kitaplar:
            messagebox.showinfo("Bilgi", "Tüm kitapların barkodu zaten mevcut.")
            return
        
        onay = messagebox.askyesno("Onay", 
            f"{len(kitaplar)} kitaba barkod oluşturulacak.\n\nDevam etmek istiyor musunuz?")
        
        if not onay:
            return
        
        olusturulan = 0
        for kitap_id, kitap_ad in kitaplar:
            try:
                barkod_kodu = self.benzersiz_barkod_olustur()
                self.cursor.execute("UPDATE kitaplar SET barkod=? WHERE id=?", (barkod_kodu, kitap_id))
                self.barkod_kaydet(barkod_kodu, kitap_ad)
                olusturulan += 1
            except Exception as e:
                print(f"Barkod oluşturulamadı: {kitap_ad} - {e}")
        
        self.conn.commit()
        messagebox.showinfo("Başarılı", f"{olusturulan} kitap için barkod oluşturuldu.")
        self.durum_yaz(f"{olusturulan} kitaba barkod eklendi.")
    
    # --- BARKOD OKUMA ---
    def barkod_tara(self):
        """Kamera ile barkod tara"""
        if not BARKOD_OKUMA_DESTEGI:
            messagebox.showerror("Hata", "Barkod tarama için gerekli kütüphaneleri yükleyin:\npip install opencv-python pyzbar")
            return
        
        # Kamera penceresi
        tarama_pencere = tk.Toplevel(self.root)
        tarama_pencere.title("Barkod Tarama")
        tarama_pencere.geometry("700x550")
        tarama_pencere.configure(bg="#2C3E50")
        
        tk.Label(tarama_pencere, text="📷 Barkodu kameraya gösterin", 
                 font=("Segoe UI", 14, "bold"), bg="#2C3E50", fg="white").pack(pady=10)
        
        video_label = tk.Label(tarama_pencere, bg="black")
        video_label.pack(pady=10)
        
        sonuc_label = tk.Label(tarama_pencere, text="Bekleniyor...", 
                                font=("Segoe UI", 12), bg="#2C3E50", fg="#BDC3C7")
        sonuc_label.pack(pady=10)
        
        cap = cv2.VideoCapture(0)
        running = [True]
        
        def update_frame():
            if not running[0]:
                return
            
            ret, frame = cap.read()
            if ret:
                # Barkod algıla
                barcodes = pyzbar.decode(frame)
                
                for barcode_obj in barcodes:
                    barkod_data = barcode_obj.data.decode('utf-8')
                    
                    # Çerçeve çiz
                    (x, y, w, h) = barcode_obj.rect
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    cv2.putText(frame, barkod_data, (x, y - 10), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                    
                    # Veritabanında ara
                    self.cursor.execute("SELECT id, ad FROM kitaplar WHERE barkod=?", (barkod_data,))
                    sonuc = self.cursor.fetchone()
                    
                    if sonuc:
                        sonuc_label.config(text=f"✅ Bulundu: {sonuc[1]}", fg="#2ECC71")
                        running[0] = False
                        cap.release()
                        tarama_pencere.after(1500, tarama_pencere.destroy)
                        
                        # Kitabı seç
                        for item in self.tree.get_children():
                            if self.tree.item(item)['values'][0] == sonuc[0]:
                                self.tree.selection_set(item)
                                self.tree.see(item)
                                break
                        
                        self.durum_yaz(f"Barkod ile bulundu: {sonuc[1]}")
                        return
                    else:
                        sonuc_label.config(text=f"❌ Bulunamadı: {barkod_data}", fg="#E74C3C")
                
                # Frame'i göster
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                frame = cv2.resize(frame, (640, 400))
                img = Image.fromarray(frame)
                imgtk = ImageTk.PhotoImage(image=img)
                video_label.imgtk = imgtk
                video_label.configure(image=imgtk)
            
            if running[0]:
                tarama_pencere.after(30, update_frame)
        
        def kapat():
            running[0] = False
            cap.release()
            tarama_pencere.destroy()
        
        tarama_pencere.protocol("WM_DELETE_WINDOW", kapat)
        tk.Button(tarama_pencere, text="❌ Kapat", command=kapat,
                  font=("Segoe UI", 11), bg="#C0392B", fg="white").pack(pady=10)
        
        update_frame()
    
    def barkod_ile_ara(self):
        """Manuel barkod girişi ile arama"""
        top = tk.Toplevel(self.root)
        top.title("Barkod ile Ara")
        top.geometry("400x180")
        top.configure(bg=self.bg_color)
        
        tk.Label(top, text="Barkod Numarası:", font=("Segoe UI", 12), 
                 bg=self.bg_color).pack(pady=20)
        
        entry = ttk.Entry(top, font=("Segoe UI", 14), width=20, justify="center")
        entry.pack(pady=5)
        entry.focus_set()
        
        def ara(event=None):
            barkod = entry.get().strip()
            if not barkod:
                return
            
            self.cursor.execute("SELECT id, ad FROM kitaplar WHERE barkod=?", (barkod,))
            sonuc = self.cursor.fetchone()
            
            if sonuc:
                # Kitabı seç
                for item in self.tree.get_children():
                    if self.tree.item(item)['values'][0] == sonuc[0]:
                        self.tree.selection_set(item)
                        self.tree.see(item)
                        break
                top.destroy()
                self.durum_yaz(f"Bulundu: {sonuc[1]}")
            else:
                messagebox.showwarning("Bulunamadı", f"'{barkod}' barkodlu kitap bulunamadı.", parent=top)
        
        entry.bind("<Return>", ara)
        tk.Button(top, text="ARA", command=ara, font=("Segoe UI", 11, "bold"),
                  bg="#27AE60", fg="white", width=15).pack(pady=15)
    
    # --- PAROLA DEĞİŞTİRME ---
    def parola_degistir_penceresi(self):
        """Öğretmen ve öğrenci parolalarını değiştirme penceresi"""
        top = tk.Toplevel(self.root)
        top.title("Parola Değiştir")
        top.geometry("380x320")
        top.configure(bg="#1a1a2e")
        top.resizable(False, False)
        
        tk.Label(top, text="🔑 Parola Yönetimi", font=("Arial", 14, "bold"),
                 bg="#1a1a2e", fg="#e94560").pack(pady=15)
        
        # Öğretmen parolası
        frame1 = tk.Frame(top, bg="#1a1a2e")
        frame1.pack(fill=tk.X, padx=30, pady=10)
        tk.Label(frame1, text="Öğretmen Parolası:", bg="#1a1a2e", fg="white",
                 font=("Arial", 10), width=18, anchor="w").pack(side=tk.LEFT)
        ogretmen_entry = tk.Entry(frame1, font=("Arial", 11), width=18)
        ogretmen_entry.pack(side=tk.RIGHT)
        
        # Mevcut parolaları yükle
        self.cursor.execute("SELECT parola FROM parolalar WHERE tip='ogretmen'")
        result = self.cursor.fetchone()
        if result:
            ogretmen_entry.insert(0, result[0])
        
        # Öğrenci parolası
        frame2 = tk.Frame(top, bg="#1a1a2e")
        frame2.pack(fill=tk.X, padx=30, pady=10)
        tk.Label(frame2, text="Öğrenci Parolası:", bg="#1a1a2e", fg="white",
                 font=("Arial", 10), width=18, anchor="w").pack(side=tk.LEFT)
        ogrenci_entry = tk.Entry(frame2, font=("Arial", 11), width=18)
        ogrenci_entry.pack(side=tk.RIGHT)
        
        self.cursor.execute("SELECT parola FROM parolalar WHERE tip='ogrenci'")
        result = self.cursor.fetchone()
        if result:
            ogrenci_entry.insert(0, result[0])
        
        def kaydet():
            yeni_ogretmen = ogretmen_entry.get().strip()
            yeni_ogrenci = ogrenci_entry.get().strip()
            
            if not yeni_ogretmen or not yeni_ogrenci:
                messagebox.showerror("Hata", "Parolalar boş bırakılamaz!", parent=top)
                return
            
            if len(yeni_ogretmen) < 4 or len(yeni_ogrenci) < 4:
                messagebox.showerror("Hata", "Parolalar en az 4 karakter olmalı!", parent=top)
                return
            
            self.cursor.execute("UPDATE parolalar SET parola=? WHERE tip='ogretmen'", (yeni_ogretmen,))
            self.cursor.execute("UPDATE parolalar SET parola=? WHERE tip='ogrenci'", (yeni_ogrenci,))
            self.conn.commit()
            
            messagebox.showinfo("Başarılı", "Parolalar güncellendi!", parent=top)
            top.destroy()
        
        tk.Button(top, text="💾 KAYDET", command=kaydet, font=("Arial", 11, "bold"),
                  bg="#e94560", fg="white", width=15, bd=0, cursor="hand2").pack(pady=25)
        
        tk.Label(top, text="Not: Değişiklikler bir sonraki girişte\ngeçerli olacaktır.", 
                 bg="#1a1a2e", fg="#666", font=("Arial", 9)).pack()
    
    # --- ÖĞRENCİ EXCEL IMPORT ---
    def ogrenci_excel_import(self):
        """Excel dosyasından öğrenci aktarma"""
        if not EXCEL_DESTEGI:
            messagebox.showerror("Hata", "Excel desteği için 'openpyxl' yükleyin")
            return
        
        dosya = filedialog.askopenfilename(
            title="Öğrenci Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )
        
        if not dosya:
            return
        
        try:
            wb = load_workbook(dosya)
            ws = wb.active
            
            # Başlıkları bul
            basliklar = []
            for cell in ws[1]:
                basliklar.append(str(cell.value).lower().strip() if cell.value else "")
            
            # Sütun eşleştirme
            sutun_map = {'okul_no': None, 'ad_soyad': None, 'sinif': None}
            
            for i, baslik in enumerate(basliklar):
                if 'no' in baslik or 'numara' in baslik:
                    sutun_map['okul_no'] = i
                elif 'ad' in baslik or 'isim' in baslik or 'soyad' in baslik:
                    if sutun_map['ad_soyad'] is None:
                        sutun_map['ad_soyad'] = i
                elif 'sınıf' in baslik or 'sinif' in baslik:
                    sutun_map['sinif'] = i
            
            # Verileri oku
            ogrenciler = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                ogrenci = {
                    'okul_no': str(row[sutun_map['okul_no']]).strip() if sutun_map['okul_no'] is not None and row[sutun_map['okul_no']] else "",
                    'ad_soyad': str(row[sutun_map['ad_soyad']]).strip() if sutun_map['ad_soyad'] is not None and row[sutun_map['ad_soyad']] else "",
                    'sinif': str(row[sutun_map['sinif']]).strip() if sutun_map['sinif'] is not None and row[sutun_map['sinif']] else ""
                }
                
                if ogrenci['okul_no'] and ogrenci['ad_soyad']:
                    ogrenciler.append(ogrenci)
            
            if not ogrenciler:
                messagebox.showwarning("Uyarı", "Excel dosyasında geçerli öğrenci bulunamadı.")
                return
            
            onay = messagebox.askyesno("Onay", 
                f"{len(ogrenciler)} öğrenci bulundu.\n\nBu öğrencileri veritabanına eklemek istiyor musunuz?")
            
            if onay:
                eklenen = 0
                guncellenen = 0
                for ogr in ogrenciler:
                    try:
                        self.cursor.execute(
                            "INSERT OR REPLACE INTO ogrenciler (okul_no, ad_soyad, sinif) VALUES (?,?,?)",
                            (ogr['okul_no'], ogr['ad_soyad'], ogr['sinif'])
                        )
                        eklenen += 1
                    except Exception as e:
                        print(f"Öğrenci eklenemedi: {ogr['ad_soyad']} - {e}")
                
                self.conn.commit()
                messagebox.showinfo("Başarılı", f"{eklenen} öğrenci aktarıldı.")
                self.durum_yaz(f"Excel'den {eklenen} öğrenci aktarıldı.")
                
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası okunamadı:\n{e}")
    
    # --- ÖĞRENCİ LİSTESİ ---
    def ogrenci_listesi_penceresi(self):
        """Öğrenci listesi ve kitap takip penceresi"""
        top = tk.Toplevel(self.root)
        top.title("Öğrenci Listesi")
        top.geometry("900x600")
        top.configure(bg="#1a1a2e")
        
        # Başlık
        header = tk.Frame(top, bg="#1a1a2e")
        header.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(header, text="👥 Öğrenci Listesi", font=("Arial", 16, "bold"),
                 bg="#1a1a2e", fg="#e94560").pack(side=tk.LEFT)
        
        # Arama
        search_frame = tk.Frame(header, bg="#1a1a2e")
        search_frame.pack(side=tk.RIGHT)
        
        tk.Label(search_frame, text="🔍 Ara:", bg="#1a1a2e", fg="white",
                 font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        
        search_entry = tk.Entry(search_frame, font=("Arial", 11), width=25)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # Tablo çerçevesi
        table_frame = tk.Frame(top, bg="#16213e")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Treeview
        columns = ("OkulNo", "AdSoyad", "Sinif", "AldigiKitap")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        tree.heading("OkulNo", text="Okul No")
        tree.heading("AdSoyad", text="Ad Soyad")
        tree.heading("Sinif", text="Sınıf")
        tree.heading("AldigiKitap", text="Aldığı Kitap")
        
        tree.column("OkulNo", width=100, anchor="center")
        tree.column("AdSoyad", width=200)
        tree.column("Sinif", width=80, anchor="center")
        tree.column("AldigiKitap", width=300)
        
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Satır renkleri
        tree.tag_configure('kitapli', background='#FDEBD0')
        tree.tag_configure('normal', background='#ffffff')
        
        def listeyi_yukle(arama=""):
            tree.delete(*tree.get_children())
            
            if arama:
                query = """
                    SELECT o.okul_no, o.ad_soyad, o.sinif, k.ad 
                    FROM ogrenciler o 
                    LEFT JOIN odunc_alanlar oa ON o.okul_no = oa.ogrenci_no 
                    LEFT JOIN kitaplar k ON oa.kitap_id = k.id
                    WHERE o.okul_no LIKE ? OR o.ad_soyad LIKE ? OR o.sinif LIKE ?
                    ORDER BY o.sinif, o.ad_soyad
                """
                self.cursor.execute(query, (f"%{arama}%", f"%{arama}%", f"%{arama}%"))
            else:
                query = """
                    SELECT o.okul_no, o.ad_soyad, o.sinif, k.ad 
                    FROM ogrenciler o 
                    LEFT JOIN odunc_alanlar oa ON o.okul_no = oa.ogrenci_no 
                    LEFT JOIN kitaplar k ON oa.kitap_id = k.id
                    ORDER BY o.sinif, o.ad_soyad
                """
                self.cursor.execute(query)
            
            for row in self.cursor.fetchall():
                kitap = row[3] if row[3] else "-"
                tag = 'kitapli' if row[3] else 'normal'
                tree.insert("", tk.END, values=(row[0], row[1], row[2], kitap), tags=(tag,))
        
        def arama_yap(event=None):
            listeyi_yukle(search_entry.get())
        
        search_entry.bind("<KeyRelease>", arama_yap)
        
        # İstatistik
        stat_frame = tk.Frame(top, bg="#1a1a2e")
        stat_frame.pack(fill=tk.X, padx=20, pady=10)
        
        self.cursor.execute("SELECT COUNT(*) FROM ogrenciler")
        toplam = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(DISTINCT ogrenci_no) FROM odunc_alanlar")
        kitapli = self.cursor.fetchone()[0]
        
        tk.Label(stat_frame, text=f"Toplam: {toplam} öğrenci | Kitap alanlar: {kitapli}",
                 bg="#1a1a2e", fg="#aaa", font=("Arial", 10)).pack()
        
        # İlk yükleme
        listeyi_yukle()
    
    # --- ÇIKIŞ YAP ---
    def cikis_yap(self):
        """Oturumu kapat ve giriş ekranına dön"""
        if messagebox.askyesno("Çıkış", "Oturumu kapatmak istiyor musunuz?"):
            self.root.destroy()
            if CTK_DESTEGI:
                yeni_root = ctk.CTk()
            else:
                yeni_root = tk.Tk()
            GirisEkrani(yeni_root, lambda tip: ana_uygulama_baslat(yeni_root, tip))
            yeni_root.mainloop()


def ana_uygulama_baslat(giris_root, kullanici_tipi):
    """Giriş ekranını kapat ve ana uygulamayı başlat"""
    giris_root.destroy()
    root = tk.Tk()
    app = KutuphaneUygulamasi(root, kullanici_tipi)
    root.mainloop()


if __name__ == "__main__":
    if CTK_DESTEGI:
        root = ctk.CTk()
    else:
        root = tk.Tk()
    GirisEkrani(root, lambda tip: ana_uygulama_baslat(root, tip))
    root.mainloop()